from __future__ import annotations

import asyncio
import csv
import datetime
import hashlib
import io
import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import STORAGE_DIR
from app.core.db import AsyncSessionLocal
from app.core.events import event_bus

from app.models.models import Artifact, Asset, Evidence, ExportJob, Finding, Port, Scan, Service, Technology, URL
from app.reporting.engine import ReportEngine

logger = logging.getLogger("export_manager")


def _get_export_dir(scan_id: str) -> Path:
    p = STORAGE_DIR / "investigations" / scan_id / "exports"
    p.mkdir(parents=True, exist_ok=True)
    return p


class ExportManager:
    """Central manager for Asynchronous Investigation Export Generation (Requirement §22, §23, §24)."""

    SUPPORTED_TYPES = {
        "full_pdf": ("application/pdf", ".pdf"),
        "executive_pdf": ("application/pdf", ".pdf"),
        "technical_pdf": ("application/pdf", ".pdf"),
        "findings_csv": ("text/csv", ".csv"),
        "findings_xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
        "investigation_json": ("application/json", ".json"),
        "assets_csv": ("text/csv", ".csv"),
        "services_csv": ("text/csv", ".csv"),
        "evidence_index_json": ("application/json", ".json"),
        "artifact_manifest_json": ("application/json", ".json"),
    }

    @classmethod
    async def create_export_job(
        cls,
        scan_id: str,
        export_type: str,
        db: AsyncSession,
    ) -> ExportJob:
        if export_type not in cls.SUPPORTED_TYPES:
            raise ValueError(f"Unsupported export type '{export_type}'. Supported: {list(cls.SUPPORTED_TYPES.keys())}")

        mime_type, ext = cls.SUPPORTED_TYPES[export_type]
        timestamp_str = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{export_type}_{scan_id}_{timestamp_str}{ext}"

        job = ExportJob(
            scan_id=scan_id,
            export_type=export_type,
            filename=filename,
            status="QUEUED",
            mime_type=mime_type,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        # Launch generation in background task
        asyncio.create_task(cls._execute_generation(job.id, scan_id, export_type, filename))
        return job

    @classmethod
    async def _execute_generation(cls, job_id: str, scan_id: str, export_type: str, filename: str) -> None:
        """Background asynchronous generator for investigation reports & dataset exports."""
        export_dir = _get_export_dir(scan_id)
        out_path = export_dir / filename

        async with AsyncSessionLocal() as db:
            job = await db.get(ExportJob, job_id)
            if not job:
                return

            job.status = "PROCESSING"
            await db.commit()

            try:
                # Gather scan data
                scan = await db.get(Scan, scan_id)
                if not scan:
                    raise RuntimeError(f"Scan '{scan_id}' not found.")

                assets = (await db.execute(select(Asset).where(Asset.scan_id == scan_id))).scalars().all()
                ports = (await db.execute(select(Port).join(Asset, Port.asset_id == Asset.id).where(Asset.scan_id == scan_id))).scalars().all()
                urls = (await db.execute(select(URL).join(Asset, URL.asset_id == Asset.id).where(Asset.scan_id == scan_id))).scalars().all()
                techs = (await db.execute(select(Technology).join(Asset, Technology.asset_id == Asset.id).where(Asset.scan_id == scan_id))).scalars().all()
                findings = (await db.execute(select(Finding).where(Finding.scan_id == scan_id).order_by(desc(Finding.first_seen)))).scalars().all()
                evidence_items = (await db.execute(select(Evidence).where(Evidence.scan_id == scan_id))).scalars().all()
                artifacts = (await db.execute(select(Artifact).where(Artifact.scan_id == scan_id))).scalars().all()

                findings_dicts = [f.to_dict() if hasattr(f, "to_dict") else {
                    "id": f.id,
                    "title": f.title,
                    "finding_code": f.finding_code,
                    "severity": f.severity,
                    "confidence": f.confidence,
                    "status": f.status,
                    "description": f.description,
                    "technical_details": f.technical_details,
                    "remediation": f.remediation,
                    "impact": f.impact,
                    "cwe_id": f.cwe_id,
                    "cve_id": f.cve_id,
                    "cvss_score": f.cvss_score,
                    "location": getattr(f, "location", "") or getattr(f, "technical_details", ""),
                    "poc": getattr(f, "proof_curl", "") or (f.evidence.get("curl") if isinstance(f.evidence, dict) else ""),
                    "created_at": f.first_seen.isoformat() if f.first_seen else None,
                } for f in findings]


                assets_dicts = [{"id": a.id, "hostname": a.hostname, "ip": a.ip, "asset_type": a.asset_type, "status": a.status} for a in assets]
                ports_dicts = [{"id": p.id, "port": getattr(p, "port_number", getattr(p, "port", None)), "service": p.service, "protocol": p.protocol} for p in ports]
                tech_dicts = [{"name": t.name, "version": t.version, "category": t.category} for t in techs]

                stats = {
                    "total_assets": len(assets),
                    "total_ports": len(ports),
                    "total_urls": len(urls),
                    "total_technologies": len(techs),
                    "total_findings": len(findings),
                    "total_artifacts": len(artifacts),
                }

                # Generate target payload
                file_bytes: bytes = b""

                if export_type in ("full_pdf", "executive_pdf", "technical_pdf"):
                    pdf_io = ReportEngine.generate_pdf(
                        scan_id=scan_id,
                        target=scan.root_domain,
                        stats=stats,
                        findings=findings_dicts,
                        assets=assets_dicts,
                        ports=ports_dicts,
                        technologies=tech_dicts,
                        operator="Autonomous Adversary Engine",
                        view_perspective="customer",
                    )
                    file_bytes = pdf_io.getvalue()

                elif export_type == "findings_csv":
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow([
                        "Finding Code", "Severity", "Title", "Status", "Confidence",
                        "Target Host / Location", "CWE", "CVE", "CVSS Score",
                        "Impact", "Description", "Technical Details", "Remediation", "PoC"
                    ])
                    for f in findings:
                        writer.writerow([
                            f.finding_code or f.id,
                            f.severity,
                            f.title,
                            f.status,
                            f.confidence,
                            getattr(f, "location", "") or getattr(f, "technical_details", "") or scan.root_domain,
                            f.cwe_id or "",
                            f.cve_id or "",
                            f.cvss_score or "",
                            f.impact or "",
                            f.description or "",
                            f.technical_details or "",
                            f.remediation or "",
                            getattr(f, "proof_curl", "") or (f.evidence.get("curl") if isinstance(f.evidence, dict) else ""),
                        ])
                    file_bytes = output.getvalue().encode("utf-8")

                elif export_type == "findings_xlsx":
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Security Findings"

                        headers = ["Code", "Severity", "Title", "Status", "Confidence", "Location", "CWE", "CVE", "CVSS", "Remediation"]
                        ws.append(headers)

                        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        for f in findings:
                            ws.append([
                                f.finding_code or f.id,
                                f.severity,
                                f.title,
                                f.status,
                                f.confidence,
                                getattr(f, "location", "") or getattr(f, "technical_details", "") or scan.root_domain,
                                f.cwe_id or "",
                                f.cve_id or "",
                                f.cvss_score or "",
                                f.remediation or "",
                            ])

                        # Assets sheet
                        ws_assets = wb.create_sheet(title="Assets Inventory")
                        ws_assets.append(["Hostname", "IP Address", "Type", "Status"])
                        for a in assets:
                            ws_assets.append([a.hostname or "", a.ip or "", a.asset_type, a.status])

                        # Services sheet
                        ws_services = wb.create_sheet(title="Discovered Services")
                        ws_services.append(["Host", "Port", "Protocol", "Service", "Product", "Version"])
                        for p in ports:
                            p_num = getattr(p, "port_number", getattr(p, "port", None)) or 0
                            ws_services.append([scan.root_domain, p_num, p.protocol, p.service or "", getattr(p, "product", ""), getattr(p, "version", "")])

                        excel_out = io.BytesIO()
                        wb.save(excel_out)
                        file_bytes = excel_out.getvalue()
                    except ImportError:
                        # Fallback to CSV if openpyxl not installed
                        output = io.StringIO()
                        writer = csv.writer(output)
                        writer.writerow(["Finding Code", "Severity", "Title", "Status", "Confidence", "Location", "CWE", "CVE", "CVSS", "Remediation"])
                        for f in findings:
                            writer.writerow([f.finding_code or f.id, f.severity, f.title, f.status, f.confidence, getattr(f, "location", ""), f.cwe_id, f.cve_id, f.cvss_score, f.remediation])
                        file_bytes = output.getvalue().encode("utf-8")

                elif export_type == "investigation_json":
                    inv_data = {
                        "investigation_id": scan_id,
                        "target": scan.root_domain,
                        "status": scan.status,
                        "statistics": stats,
                        "findings": findings_dicts,
                        "assets": assets_dicts,
                        "services": ports_dicts,
                        "technologies": tech_dicts,
                        "exported_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                    }
                    file_bytes = json.dumps(inv_data, indent=2).encode("utf-8")

                elif export_type == "assets_csv":
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(["ID", "Hostname", "IP Address", "Asset Type", "Status"])
                    for a in assets:
                        writer.writerow([a.id, a.hostname or "", a.ip or "", a.asset_type, a.status])
                    file_bytes = output.getvalue().encode("utf-8")

                elif export_type == "services_csv":
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(["Port", "Protocol", "Service", "Product", "Version", "Banner", "TLS Enabled"])
                    for p in ports:
                        p_num = getattr(p, "port_number", getattr(p, "port", None)) or 0
                        writer.writerow([
                            p_num,
                            p.protocol or "tcp",
                            p.service or "unknown",
                            getattr(p, "product", ""),
                            getattr(p, "version", ""),
                            getattr(p, "banner", ""),
                            getattr(p, "is_tls", False),
                        ])
                    file_bytes = output.getvalue().encode("utf-8")

                elif export_type == "evidence_index_json":
                    ev_index = []
                    for ev in evidence_items:
                        ev_index.append({
                            "evidence_id": ev.id,
                            "finding_id": getattr(ev, "finding_id", None) or (ev.data.get("finding_id") if isinstance(ev.data, dict) else None),
                            "evidence_type": getattr(ev, "evidence_type", "http_interaction"),
                            "sha256_hash": getattr(ev, "sha256_hash", None),
                            "request_data": getattr(ev, "request_dump", None) or getattr(ev, "request", None) or (ev.data.get("request_headers") if isinstance(ev.data, dict) else None),
                            "response_data": getattr(ev, "response_dump", None) or getattr(ev, "response", None) or (ev.data.get("response_headers") if isinstance(ev.data, dict) else None),
                            "timestamp": ev.created_at.isoformat() if ev.created_at else None,
                        })
                    file_bytes = json.dumps({"scan_id": scan_id, "total_evidence": len(ev_index), "evidence": ev_index}, indent=2).encode("utf-8")

                elif export_type == "artifact_manifest_json":
                    art_manifest = []
                    for art in artifacts:
                        art_manifest.append({
                            "id": art.id,
                            "filename": art.filename,
                            "file_type": art.file_type,
                            "classification": art.classification,
                            "category": art.category,
                            "record_count": art.record_count,
                            "size_bytes": art.size_bytes,
                            "sha256_hash": art.sha256_hash,
                            "source_url": art.source,
                            "is_redacted": art.is_redacted,
                            "created_at": art.created_at.isoformat() if art.created_at else None,
                        })
                    file_bytes = json.dumps({"scan_id": scan_id, "total_artifacts": len(art_manifest), "artifacts": art_manifest}, indent=2).encode("utf-8")

                # Write to filesystem
                out_path.write_bytes(file_bytes)
                sha256 = hashlib.sha256(file_bytes).hexdigest()

                job.status = "COMPLETED"
                job.file_path = str(out_path)
                job.file_size = len(file_bytes)
                job.sha256_hash = sha256
                job.completed_at = datetime.datetime.now(datetime.timezone.utc)
                await db.commit()

                # Publish event
                await event_bus.publish({
                    "scan_id": scan_id,
                    "event_type": "export.completed",
                    "data": {
                        "job_id": job.id,
                        "export_type": export_type,
                        "filename": filename,
                        "file_size": len(file_bytes),
                        "sha256": sha256,
                    },
                    "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                })

            except Exception as err:
                logger.error("Async export generation failed for %s (%s): %s", scan_id, export_type, err, exc_info=True)
                job.status = "FAILED"
                job.error_message = str(err)
                await db.commit()
