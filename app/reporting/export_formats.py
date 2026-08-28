"""Pure export rendering, run in a worker thread with sanitized snapshots."""
import csv
import io
import json

from app.reporting.engine import ReportEngine
from app.reporting.redaction import RedactionEngine


def spreadsheet_cell(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    if isinstance(value, str):
        value = RedactionEngine.redact_text(value)
        # Neutralize spreadsheet formulas, including leading invisible whitespace.
        if value.lstrip().startswith(("=", "+", "-", "@")) or value.startswith(("\t", "\r", "\n")):
            value = "'" + value
    return value


def render_export(export_type: str, snapshot: dict) -> bytes:
    data = RedactionEngine.redact_dict(snapshot)
    findings, assets, services = data["findings"], data["assets"], data["services"]
    if export_type.endswith("_pdf"):
        return ReportEngine.generate_pdf(
            scan_id=data["investigation_id"], target=data["target"], stats=data["statistics"],
            findings=findings, assets=assets, ports=services, technologies=data["technologies"],
            report_type=export_type.removesuffix("_pdf"))
    if export_type in {"investigation_json", "evidence_index_json", "artifact_manifest_json"}:
        payload = data
        if export_type == "evidence_index_json":
            payload = {"scan_id": data["investigation_id"], "evidence": data["evidence"], "total_evidence": len(data["evidence"])}
        if export_type == "artifact_manifest_json":
            payload = {"scan_id": data["investigation_id"], "artifacts": data["artifacts"], "total_artifacts": len(data["artifacts"])}
        return json.dumps(payload, indent=2, ensure_ascii=False, default=str).encode("utf-8")
    finding_keys = ["finding_code", "severity", "title", "status", "confidence", "evidence_level",
                    "location", "asset_hostname", "cwe_id", "cve_id", "cvss_score", "impact",
                    "description", "actual_result", "expected_result", "remediation", "poc",
                    "report_quality", "first_seen", "last_seen"]
    asset_keys = ["id", "hostname", "ip", "asset_type", "status"]
    service_keys = ["asset_id", "host", "port", "protocol", "service", "banner"]
    records, keys = (assets, asset_keys) if export_type == "assets_csv" else ((services, service_keys) if export_type == "services_csv" else (findings, finding_keys))

    def rows(items, fields):
        yield fields
        for item in items:
            yield [spreadsheet_cell(item.get(key)) for key in fields]

    if export_type == "findings_xlsx":
        # Missing dependency is an explicit failed job, never CSV with an XLSX extension.
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        for title, items, fields in (("Findings", findings, finding_keys), ("Assets", assets, asset_keys), ("Services", services, service_keys)):
            sheet = workbook.create_sheet(title)
            for row in rows(items, fields):
                sheet.append(row)
            for row in sheet:
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.data_type = "s"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E293B")
                sheet.column_dimensions[cell.column_letter].width = 28
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerows(rows(records, keys))
    return output.getvalue().encode("utf-8-sig")
